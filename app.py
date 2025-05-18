# --- Standard Library ---
import os
import csv
import io
import json
import random
import zipfile
from datetime import datetime, timedelta
from io import BytesIO, StringIO

import os
import json
import random
import pandas as pd
import openpyxl
from io import BytesIO, StringIO
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from flask import send_file
from fpdf import FPDF
from io import BytesIO
import sqlite3
from student_profile_dashboard import course_modules, module_meta

# --- Third-Party Libraries ---
import pandas as pd
import openpyxl
import sqlite3
import pdfkit
from fpdf import FPDF
from faker import Faker
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


from flask import (
    Flask, render_template, redirect, url_for, request, flash, session,
    jsonify, send_file, make_response
)
from flask_sqlalchemy import SQLAlchemy
from flask_login import (
    LoginManager, UserMixin, login_user, login_required,
    logout_user, current_user
)
from flask_bcrypt import Bcrypt

# --- Dash Libraries ---
import dash_bootstrap_components as dbc
import dash_html_components as html
import dash_core_components as dcc

# --- Local Modules ---
from models import db, User
from data_persistence import load_data
from report_data_loader import fetch_module_data
from student_data_loader import fetch_student_dashboard
from student_profile_dashboard import course_modules, module_meta
from dashboard_admin import fetch_course_data
from data_science_dashboard import raw_data as ds_raw
from software_engineering_dashboard import raw_data as se_raw
from dashboard_big_data import big_data_students
from mlops_dashboard import raw_data as mlops_raw
from data_ethics_dashboard import raw_data as ethics_raw
from software_testing_dashboard import raw_data as testing_raw
from cloud_engineering_dashboard import raw_data as cloud_raw
from dashboard_web_systems import web_systems_students

path_wkhtmltopdf = r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe"
config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)


# --- Log in Logic ---
app = Flask(__name__)
app.config['SECRET_KEY'] = 'supersecretkey'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)
bcrypt = Bcrypt(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'  

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@app.route('/')
def index():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        if user and bcrypt.check_password_hash(user.password, password):
            login_user(user)
            flash("Login Successful!", "success")
            if user.role == "student":
                return redirect(url_for('home_student'))
            elif user.role == "lecturer":
                return redirect(url_for('home_lecturer'))
            else:
                return redirect(url_for('home_admin'))
        else:
            flash("Invalid username or password", "danger")
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    session.pop('_flashes', None) 
    flash("You have been logged out.", "info")
    return redirect(url_for('login'))

# --- Student side logic ---

@app.route('/home_student')
@login_required
def home_student():
    return render_template('home_student.html', user=current_user)


@app.route('/student_profile')
@login_required
def student_profile():
    if current_user.role != "student":
        flash("Unauthorized Access", "danger")
        return redirect(url_for('index'))
    return render_template('student_profile.html', user=current_user)

@app.route('/student/modules')
@login_required
def student_modules():
    return render_template('student_modules.html')

@app.route('/data_structures_dashboard')
@login_required
def data_structures_dashboard():
    return render_template('data_structures_dashboard.html')

@app.route('/webdev_dashboard')
@login_required
def webdev_dashboard():
    return render_template('webdev_dashboard.html')

@app.route('/ai_dashboard')
@login_required
def ai_dashboard():
    return render_template('ai_dashboard.html')

@app.route('/db_dashboard')
@login_required
def db_dashboard():
    return render_template('db_dashboard.html')

@app.route('/cybersecurity_dashboard')
@login_required
def cybersecurity_dashboard():
    return render_template('cybersecurity_dashboard.html')


@app.route('/attendance_dashboard')
@login_required
def attendance_dashboard():
    return render_template('attendance_dashboard.html')



@app.route('/export')
def export():
    return render_template('export.html')  


# EXPORT for STUDENT-SIDE MODULES #
def generate_module_report_from_dict(module_data, file_type):
    
    safe_name = module_data['module_name'].replace(" ", "_")  
    file_path = f"{safe_name}_Report.{file_type}"
    metrics = module_data["metrics"]
    assignments = module_data["assignments"]

    if file_type == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = module_data['module_name']
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 20

        ws.merge_cells("A1:B1")
        ws["A1"] = f"MODULE: {module_data['module_name']}"
        ws["A1"].font = Font(bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:B2")
        ws["A2"] = f"Module Code: {module_data['module_code']}"
        ws["A2"].alignment = Alignment(horizontal="center")

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for i, (metric, value) in enumerate(metrics):
            ws[f"A{4 + i}"] = metric
            ws[f"B{4 + i}"] = value
            ws[f"A{4 + i}"].border = thin_border
            ws[f"B{4 + i}"].border = thin_border

        ws["A8"] = "ASSESSMENT"
        ws["B8"] = "SCORE %"
        ws["A8"].font = Font(bold=True)
        ws["B8"].font = Font(bold=True)

        for i, (title, score) in enumerate(assignments):
            ws[f"A{9 + i}"] = title
            ws[f"B{9 + i}"] = score if score is not None else ""

        table = Table(displayName="ScoresTable", ref=f"A8:B{8 + len(assignments)}")
        style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        table.tableStyleInfo = style
        ws.add_table(table)

        wb.save(file_path)
        return send_file(file_path, as_attachment=True, download_name=f"{safe_name}_Report.xlsx")

    elif file_type == "csv":
        with open(file_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([f"MODULE: {module_data['module_name']}"])
            writer.writerow([f"Module Code: {module_data['module_code']}"])
            writer.writerow([])
            writer.writerow(["Metric", "Value"])
            for metric, value in metrics:
                writer.writerow([metric, value])
            writer.writerow([])
            writer.writerow(["ASSESSMENT", "SCORE %"])
            for title, score in assignments:
                writer.writerow([title, score if score is not None else ""])

        return send_file(file_path, as_attachment=True, download_name=f"{safe_name}_Report.csv")

@app.route('/download_excel_ai')
@login_required
def download_excel_ai():
    module = fetch_module_data("CS203")
    if not module:
        return "Module not found", 404
    return generate_module_report_from_dict(module, "xlsx")

@app.route('/download_csv_ai')
@login_required
def download_csv_ai():
    module = fetch_module_data("CS203")
    if not module:
        return "Module not found", 404
    return generate_module_report_from_dict(module, "csv")

@app.route('/download_excel_webdev')
@login_required
def download_excel_webdev():
    module = fetch_module_data("CS202")
    if not module:
        return "Module not found", 404
    return generate_module_report_from_dict(module, "xlsx")

@app.route('/download_csv_webdev')
@login_required
def download_csv_webdev():
    module = fetch_module_data("CS202")
    if not module:
        return "Module not found", 404
    return generate_module_report_from_dict(module, "csv")

@app.route('/download_excel_db')
@login_required
def download_excel_db():
    module = fetch_module_data("CS204")
    if not module:
        return "Module not found", 404
    return generate_module_report_from_dict(module, "xlsx")

@app.route('/download_csv_db')
@login_required
def download_csv_db():
    module = fetch_module_data("CS204")
    if not module:
        return "Module not found", 404
    return generate_module_report_from_dict(module, "csv")

@app.route('/download_excel_cybersecurity')
@login_required
def download_excel_cybersecurity():
    module = fetch_module_data("CS205")
    if not module:
        return "Module not found", 404
    return generate_module_report_from_dict(module, "xlsx")

@app.route('/download_csv_cybersecurity')
@login_required
def download_csv_cybersecurity():
    module = fetch_module_data("CS205")
    if not module:
        return "Module not found", 404
    return generate_module_report_from_dict(module, "csv")

@app.route('/download_excel_dsa')
@login_required
def download_excel_dsa():
    module = fetch_module_data("CS201")
    if not module:
        return "Module not found", 404
    return generate_module_report_from_dict(module, "xlsx")

@app.route('/download_csv_dsa')
@login_required
def download_csv_dsa():
    module = fetch_module_data("CS201")
    if not module:
        return "Module not found", 404
    return generate_module_report_from_dict(module, "csv")



def generate_student_dashboard_report(data, file_type):
    file_path = f"Student_Performance_Report.{file_type}"

    grades = data["grades"]
    assignments = data["assignments"]
    exams = data["exams"]
    deadlines = data["deadlines"]

    if file_type == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Student Dashboard"

        # Styles
        header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        completed_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        not_completed_fill = PatternFill(start_color="FF5050", end_color="FF5050", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Header
        ws.merge_cells("A1:D1")
        ws["A1"] = "STUDENT PERFORMANCE DASHBOARD"
        ws["A1"].font = Font(bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:D2")
        ws["A2"] = f"Student: {data['student_name']}"
        ws["A2"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A3:D3")
        ws["A3"] = f"Student ID: {data['student_id']}"
        ws["A3"].alignment = Alignment(horizontal="center")

        ws["A5"] = "Average Attendance"
        ws["B5"] = data["attendance"]

        # Grades Table
        ws.merge_cells("A7:D7")
        ws["A7"] = "Grades by Module"
        ws["A7"].fill = header_fill
        ws["A7"].alignment = Alignment(horizontal="center")

        ws.append(["Module Name", "Grade %"])
        for row in grades:
            ws.append(row)

        for row in ws.iter_rows(min_row=9, max_row=9 + len(grades), min_col=1, max_col=2):
            for cell in row:
                cell.border = border

        # Assignments
        ws.append([""])
        ws.merge_cells(f"A{ws.max_row + 1}:D{ws.max_row + 1}")
        ws.cell(row=ws.max_row, column=1).value = "Assignments"
        ws.cell(row=ws.max_row, column=1).fill = header_fill
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(horizontal="center")

        ws.append(["Module Name", "Assignment Name", "Score", "Status"])
        for row in assignments:
            ws.append(row)

        for row in ws.iter_rows(min_row=ws.max_row - len(assignments), max_row=ws.max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = border
                if cell.column == 4:
                    if cell.value == "Completed":
                        cell.fill = completed_fill
                    elif cell.value in ["Not Completed", "Not Started"]:
                        cell.fill = not_completed_fill

        # Exams
        ws.append([""])
        ws.merge_cells(f"A{ws.max_row + 1}:D{ws.max_row + 1}")
        ws.cell(row=ws.max_row, column=1).value = "Exams"
        ws.cell(row=ws.max_row, column=1).fill = header_fill
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(horizontal="center")

        ws.append(["Module Name", "Exam Name", "Score", "Status"])
        for row in exams:
            ws.append(row)

        for row in ws.iter_rows(min_row=ws.max_row - len(exams), max_row=ws.max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = border
                if cell.column == 4 and cell.value == "Not Completed":
                    cell.fill = not_completed_fill

        # Deadlines
        ws.append([""])
        ws.merge_cells(f"A{ws.max_row + 1}:D{ws.max_row + 1}")
        ws.cell(row=ws.max_row, column=1).value = "Upcoming Deadlines"
        ws.cell(row=ws.max_row, column=1).fill = header_fill
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(horizontal="center")

        ws.append(["Module Name", "Assignment Name", "Deadline Date", "Days Left"])
        for row in deadlines:
            ws.append(row)

        for row in ws.iter_rows(min_row=ws.max_row - len(deadlines), max_row=ws.max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = border

        wb.save(file_path)
        return send_file(file_path, as_attachment=True, download_name=file_path)

    # CSV version
    elif file_type == "csv":
        with open(file_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["STUDENT PERFORMANCE DASHBOARD"])
            writer.writerow([f"Student: {data['student_name']}"])
            writer.writerow([f"Student ID: {data['student_id']}"])
            writer.writerow([])
            writer.writerow(["Average Attendance", data["attendance"]])
            writer.writerow([])

            writer.writerow(["Module Name", "Grade %"])
            for row in grades:
                writer.writerow(row)
            writer.writerow([])

            writer.writerow(["Module Name", "Assignment Name", "Score", "Status"])
            for row in assignments:
                writer.writerow(row)
            writer.writerow([])

            writer.writerow(["Module Name", "Exam Name", "Score", "Status"])
            for row in exams:
                writer.writerow(row)
            writer.writerow([])

            writer.writerow(["Module Name", "Assignment Name", "Deadline Date", "Days Left"])
            for row in deadlines:
                writer.writerow(row)

        return send_file(file_path, as_attachment=True, download_name=file_path)


@app.route('/download_excel_student_performance')
@login_required
def download_excel_student_performance():
    data = fetch_student_dashboard()
    return generate_student_dashboard_report(data, "xlsx")

@app.route('/download_csv_student_performance')
@login_required
def download_csv_student_performance():
    data = fetch_student_dashboard()
    return generate_student_dashboard_report(data, "csv")

@app.route('/download_excel_attendance')
@login_required
def download_excel_attendance():
    conn = sqlite3.connect("reports_data.db")
    df = pd.read_sql_query("""
        SELECT m.module_name AS 'Module', r.avg_attendance AS 'Attendance (%)'
        FROM report_data r
        JOIN modules m ON r.module_code = m.module_code
    """, conn)
    conn.close()
    return export_file(df, "Attendance_Report", "xlsx")

@app.route('/download_csv_attendance')
@login_required
def download_csv_attendance():
    conn = sqlite3.connect("reports_data.db")
    df = pd.read_sql_query("""
        SELECT m.module_name AS 'Module', r.avg_attendance AS 'Attendance (%)'
        FROM report_data r
        JOIN modules m ON r.module_code = m.module_code
    """, conn)
    conn.close()
    return export_file(df, "Attendance_Report", "csv")

# --- End of Student side logic ---



# --- Lecturer side logic ---
@app.route('/home_lecturer')
@login_required
def home_lecturer():
    return render_template('home_lecturer.html', user=current_user)


@app.route('/students_overview')
@login_required
def students_overview_page():
    return render_template('students_overview.html')

@app.route("/student_attendance_insights/")
@login_required
def student_attendance_insights_page():
    return render_template("student_attendance_insights.html")




@app.route('/software_engineering_dashboard')
@login_required
def software_engineering_dashboard():
    return render_template('software_engineering_dashboard.html')

@app.route('/data_science_dashboard')
@login_required
def data_science_dashboard_page():
    return render_template('data_science_dashboard.html')

@app.route('/course_registers')
def course_registers_page():
    return render_template('course_registers.html')

@app.route('/lecturer_profile')
@login_required
def lecturer_profile():
    return render_template('lecturer_profile.html')

@app.route('/courses')
def courses():
    return render_template('courses.html') 




@app.route('/export_lecturer')
@login_required
def export_lecturer():
    return render_template('lecturer_export.html')

@app.route('/download_excel_overview_dashboard')
@login_required
def download_excel_overview_dashboard():
    return send_file("path/to/overview_dashboard.xlsx", as_attachment=True)

@app.route('/download_csv_overview_dashboard')
@login_required
def download_csv_overview_dashboard():
    return send_file("path/to/overview_dashboard.csv", as_attachment=True)

@app.route('/download_excel_software_engineering')
@login_required
def download_excel_software_engineering():
    return send_file("path/to/software_engineering.xlsx", as_attachment=True)

@app.route('/download_csv_software_engineering')
@login_required
def download_csv_software_engineering():
    return send_file("path/to/software_engineering.csv", as_attachment=True)

@app.route('/download_excel_data_science')
@login_required
def download_excel_data_science():
    return send_file("path/to/data_science.xlsx", as_attachment=True)

@app.route('/download_csv_data_science')
@login_required
def download_csv_data_science():
    return send_file("path/to/data_science.csv", as_attachment=True)

@app.route('/download_excel_students_overview')
@login_required
def download_excel_students_overview():
    return send_file("path/to/students_overview.xlsx", as_attachment=True)

@app.route('/download_csv_students_overview')
@login_required
def download_csv_students_overview():
    return send_file("path/to/students_overview.csv", as_attachment=True)

@app.route('/download_excel_student_attendance_insights')
@login_required
def download_excel_student_attendance_insights():
    return send_file("path/to/student_attendance_insights.xlsx", as_attachment=True)

@app.route('/download_csv_student_attendance_insights')
@login_required
def download_csv_student_attendance_insights():
    return send_file("path/to/student_attendance_insights.csv", as_attachment=True)

def export_file(df, filename, file_type="excel"):
    if file_type == "excel":
        file_path = f"{filename}.xlsx"
        df.to_excel(file_path, index=False)
    else:
        file_path = f"{filename}.csv"
        df.to_csv(file_path, index=False)

    return send_file(file_path, as_attachment=True)




# lecturers exports
# Function to generate Lecturer Dashboard Export
def generate_lecturer_dashboard_report(file_type="xlsx"):
    file_path = f"overview_dashboard.{file_type}"

    from course_data import software_engineering_students, data_science_students
    import pandas as pd

    def calculate_final_grade(a1, a2, exam, p1, p2):
        try:
            a1 = float(a1) if a1 is not None else 0
            a2 = float(a2) if a2 is not None else 0
            exam = float(exam) if exam is not None else 0
            grade = (a1 * 0.25) + (a2 * 0.25) + (exam * 0.5)

            if isinstance(p1, str):
                if "Lateness Penalty" in p1:
                    grade *= 0.95
                if "Word Count Penalty" in p1:
                    grade *= 0.9
            if isinstance(p2, str):
                if "Lateness Penalty" in p2:
                    grade *= 0.95
                if "Word Count Penalty" in p2:
                    grade *= 0.9

            return round(grade, 2)
        except:
            return 0

    # Load and process  data
    se_df = pd.DataFrame(software_engineering_students)
    ds_df = pd.DataFrame(data_science_students)

    se_df["Final Grade"] = se_df.apply(lambda row: calculate_final_grade(
        row.get("a1_score"),
        row.get("a2_score"),
        row.get("exam_score"),
        row.get("a1_penalty"),
        row.get("a2_penalty")
    ), axis=1)

    ds_df["Final Grade"] = ds_df.apply(lambda row: calculate_final_grade(
        row.get("a1_score"),
        row.get("a2_score"),
        row.get("exam_score"),
        row.get("a1_penalty"),
        row.get("a2_penalty")
    ), axis=1)

    # Metrics
    se_avg = se_df["Final Grade"].mean() if not se_df.empty else 0
    ds_avg = ds_df["Final Grade"].mean() if not ds_df.empty else 0
    avg_attendance = (se_df["attendance"].mean() + ds_df["attendance"].mean()) / 2 if not se_df.empty and not ds_df.empty else 0
    remaining_attendance = round(100 - avg_attendance, 2)

    metrics = [
        ["Metric", "Value"],
        ["Average Attendance (%)", round(avg_attendance, 2)],
        ["Remaining Attendance (%)", remaining_attendance],
        ["Software Engineering Avg Score (%)", round(se_avg, 2)],
        ["Data Science Avg Score (%)", round(ds_avg, 2)]
    ]
    assignments = [
    ["Assignment", "Completed", "With Penalty", "Not Completed", "Absent"],
    ["SE - Assignment 1", 42, 8, 5, 3],
    ["SE - Assignment 2", 40, 10, 6, 2],
    ["DS - Assignment 1", 38, 9, 4, 6],
    ["DS - Assignment 2", 41, 7, 3, 6]
]

   


    # Export Logic
    if file_type == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lecturer Overview Report"

        header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        # 📌 Title
        ws.merge_cells("A1:B1")
        ws["A1"] = "Lecturer Dashboard Overview"
        ws["A1"].font = Font(bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        # 📌 Metrics Table
        for row_idx, row in enumerate(metrics, start=3):
            ws.append(row)
            ws[f"A{row_idx}"].border = border
            ws[f"B{row_idx}"].border = border
            if row_idx == 3:
                ws[f"A{row_idx}"].fill = header_fill
                ws[f"B{row_idx}"].fill = header_fill

        # 📌 Assignments Table
        ws.append([""])  # Spacer
        start_row = len(metrics) + 5
        for row_idx, row in enumerate(assignments, start=start_row):
            ws.append(row)
            for col in range(1, len(row) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = border
                if row_idx == start_row:
                    cell.fill = header_fill

        wb.save(file_path)
        return file_path

    elif file_type == "csv":
        df_metrics = pd.DataFrame(metrics[1:], columns=metrics[0])
        df_assignments = pd.DataFrame(assignments[1:], columns=assignments[0])

        with open(file_path, "w") as f:
            df_metrics.to_csv(f, index=False)
            f.write("\n")
            df_assignments.to_csv(f, index=False)

        return file_path

    return None


# ✅ Flask Routes to Export Lecturer's Performance Dashboard
@app.route('/download_excel_lecturer_dashboard')
@login_required
def download_excel_lecturer_dashboard():
    file_path = generate_lecturer_dashboard_report("xlsx")
    return send_file(file_path, as_attachment=True, download_name="Lecturer_Overview_Dashboard.xlsx")

@app.route('/download_csv_lecturer_dashboard')
@login_required
def download_csv_lecturer_dashboard():
    file_path = generate_lecturer_dashboard_report("csv")
    return send_file(file_path, as_attachment=True, download_name="Lecturer_Overview_Dashboard.csv")

#  Function to generate Software Engineering Dashboard Export
def generate_software_engineering_report(file_type="excel"):
    from course_data import software_engineering_students
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    def calculate_final_grade(a1, a2, exam, p1, p2):
        try:
            a1 = float(a1) if a1 is not None else 0
            a2 = float(a2) if a2 is not None else 0
            exam = float(exam) if exam is not None else 0
            grade = (a1 * 0.25) + (a2 * 0.25) + (exam * 0.5)

            if isinstance(p1, str):
                if "Lateness Penalty" in p1:
                    grade *= 0.95
                if "Word Count Penalty" in p1:
                    grade *= 0.9
            if isinstance(p2, str):
                if "Lateness Penalty" in p2:
                    grade *= 0.95
                if "Word Count Penalty" in p2:
                    grade *= 0.9

            return round(grade, 2)
        except:
            return 0

    df = pd.DataFrame(software_engineering_students)
    df["Final Grade"] = df.apply(lambda row: calculate_final_grade(
        row.get("a1_score"),
        row.get("a2_score"),
        row.get("exam_score"),
        row.get("a1_penalty"),
        row.get("a2_penalty")
    ), axis=1)

    df["Student"] = df["name"]
    df["ID"] = df["id"]
    df["Attendance (%)"] = df["attendance"]
    df["Exam Status"] = df["exam_status"]
    df["Exam Score"] = df["exam_score"]
    df["Assignment 1 (Bugs and Fixes)"] = df["a1_score"]
    df["Assignment 1 Status"] = df["a1_status"]
    df["Assignment 1 Penalty"] = df["a1_penalty"]
    df["Assignment 2 (Software Architecture)"] = df["a2_score"]
    df["Assignment 2 Status"] = df["a2_status"]
    df["Assignment 2 Penalty"] = df["a2_penalty"]
    df["Status"] = df["status"]

    df = df[[ 
        "ID", "Student", "Final Grade", "Attendance (%)", "Exam Status", "Exam Score",
        "Assignment 1 (Bugs and Fixes)", "Assignment 1 Status", "Assignment 1 Penalty",
        "Assignment 2 (Software Architecture)", "Assignment 2 Status", "Assignment 2 Penalty",
        "Status"
    ]]

    if file_type == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Software Eng Dashboard"

        header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        title_cell = ws.cell(row=1, column=1, value="Software Engineering Dashboard Report")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        for col_num, column_title in enumerate(df.columns, start=1):
            cell = ws.cell(row=2, column=col_num, value=column_title)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(df.itertuples(index=False), start=3):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value).border = border

        with BytesIO() as b:
            wb.save(b)
            return b.getvalue()

    elif file_type == "csv":
        return df.to_csv(index=False).encode("utf-8")

    return None


@app.route('/download_excel_software_engineering_dashboard')
@login_required
def download_excel_software_engineering_dashboard():
    excel_data = generate_software_engineering_report("xlsx")
    return send_file(
        BytesIO(excel_data),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='Software_Engineering_Report.xlsx'
    )

@app.route('/download_csv_software_engineering_dashboard')
@login_required
def download_csv_software_engineering_dashboard():
    csv_data = generate_software_engineering_report("csv")
    return send_file(
        BytesIO(csv_data),
        mimetype='text/csv',
        as_attachment=True,
        download_name='Software_Engineering_Report.csv'
    )


#  Function to generate Data Science Dashboard Export
def generate_data_science_report(file_type="excel"):
    from course_data import data_science_students
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    def calculate_final_grade(a1, a2, exam, p1, p2):
        try:
            a1 = float(a1) if a1 is not None else 0
            a2 = float(a2) if a2 is not None else 0
            exam = float(exam) if exam is not None else 0
            grade = (a1 * 0.25) + (a2 * 0.25) + (exam * 0.5)

            if isinstance(p1, str):
                if "Lateness Penalty" in p1:
                    grade *= 0.95
                if "Word Count Penalty" in p1:
                    grade *= 0.9
            if isinstance(p2, str):
                if "Lateness Penalty" in p2:
                    grade *= 0.95
                if "Word Count Penalty" in p2:
                    grade *= 0.9

            return round(grade, 2)
        except:
            return 0

    df = pd.DataFrame(data_science_students)
    df["Final Grade"] = df.apply(lambda row: calculate_final_grade(
        row.get("a1_score"),
        row.get("a2_score"),
        row.get("exam_score"),
        row.get("a1_penalty"),
        row.get("a2_penalty")
    ), axis=1)

    df["Student"] = df["name"]
    df["ID"] = df["id"]
    df["Attendance (%)"] = df["attendance"]
    df["Exam Status"] = df["exam_status"]
    df["Exam Score"] = df["exam_score"]
    df["Assignment 1 (Data Analysis)"] = df["a1_score"]
    df["Assignment 1 Status"] = df["a1_status"]
    df["Assignment 1 Penalty"] = df["a1_penalty"]
    df["Assignment 2 (Machine Learning)"] = df["a2_score"]
    df["Assignment 2 Status"] = df["a2_status"]
    df["Assignment 2 Penalty"] = df["a2_penalty"]
    df["Status"] = df["status"]

    df = df[[ 
        "ID", "Student", "Final Grade", "Attendance (%)", "Exam Status", "Exam Score",
        "Assignment 1 (Data Analysis)", "Assignment 1 Status", "Assignment 1 Penalty",
        "Assignment 2 (Machine Learning)", "Assignment 2 Status", "Assignment 2 Penalty",
        "Status"
    ]]

    if file_type == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data Science Dashboard"

        header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        title_cell = ws.cell(row=1, column=1, value="Data Science Dashboard Report")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        for col_num, column_title in enumerate(df.columns, start=1):
            cell = ws.cell(row=2, column=col_num, value=column_title)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(df.itertuples(index=False), start=3):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value).border = border

        with BytesIO() as b:
            wb.save(b)
            return b.getvalue()

    elif file_type == "csv":
        return df.to_csv(index=False).encode("utf-8")

    return None



#  Flask Routes to Export Data Science Report
@app.route('/download_excel_data_science_dashboard')
@login_required
def download_excel_data_science_dashboard():
    excel_data = generate_data_science_report("xlsx")
    return send_file(
        BytesIO(excel_data),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='Data_Science_Report.xlsx'
    )

@app.route('/download_csv_data_science_dashboard')
@login_required
def download_csv_data_science_dashboard():
    csv_data = generate_data_science_report("csv")
    return send_file(
        BytesIO(csv_data),
        mimetype='text/csv',
        as_attachment=True,
        download_name='Data_Science_Report.csv'
    )



def generate_students_overview_report(file_type="xlsx"):
    try:
        se_data = load_data("software_engineering.json")
        ds_data = load_data("data_science.json")

        def flatten(data):
            flat = []
            for item in data:
                if isinstance(item, list):
                    flat.extend(item)
                elif isinstance(item, dict):
                    flat.append(item)
            return flat

        se_data = flatten(se_data)
        ds_data = flatten(ds_data)

    except Exception as e:
        print("❌ Error loading data:", e)
        return None

    for row in se_data:
        row["Course"] = "Software Engineering"
    for row in ds_data:
        row["Course"] = "Data Science"

    df_se = pd.DataFrame(se_data)
    df_ds = pd.DataFrame(ds_data)
    df = pd.concat([df_se, df_ds], ignore_index=True)

    df.rename(columns={
        "name": "Student",
        "Final Grade": "Final Grade",
        "Attendance": "Attendance",
        "attendance": "Attendance",
        "exam_score": "Exam Score",
        "id": "ID"
    }, inplace=True)

    for col in ["Final Grade", "Attendance", "Exam Score"]:
        df[col] = pd.to_numeric(df.get(col, 0), errors="coerce").fillna(0)

    # === Calculated Tables ===
    avg_grades = df.groupby("Course")["Final Grade"].mean().reset_index().rename(columns={"Final Grade": "Average Grade (%)"})
    avg_attendance = df.groupby("Course")["Attendance"].mean().reset_index().rename(columns={"Attendance": "Average Attendance (%)"})

    # ✅ At-risk = below 60
    at_risk = df[df["Final Grade"] < 60][["Student", "Course", "Final Grade"]].rename(columns={"Final Grade": "Grade"})

    # ✅ Top 5 performers per course
    top_se = df_se.sort_values("Final Grade", ascending=False).head(5)[["Student", "Course", "Final Grade"]]
    top_ds = df_ds.sort_values("Final Grade", ascending=False).head(5)[["Student", "Course", "Final Grade"]]
    top = pd.concat([top_se, top_ds]).rename(columns={"Final Grade": "Grade"})

    # ✅ Grade Distribution Table
    def get_grade_band(score):
        if score >= 70:
            return "100-70%"
        elif score >= 60:
            return "70-60%"
        elif score >= 50:
            return "60-50%"
        else:
            return "50-40%"

    df["Grade Range"] = df["Final Grade"].apply(get_grade_band)
    grade_dist = df.groupby(["Course", "Grade Range"]).size().reset_index(name="Count")

    
    performance_comparison = pd.DataFrame({
        "Metric": ["Assignments", "Exams"],
        "Software Engineering": [78, 85],
        "Data Science": [74, 82]
    })

    # === Excel Export ===
    if file_type == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students Overview"

        title_font = Font(size=14, bold=True)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))

        def write_table(start_row, start_col, title, df_table):
            col_offset = start_col
            row = start_row
            ws.merge_cells(start_row=row, start_column=col_offset, end_row=row, end_column=col_offset + len(df_table.columns) - 1)
            title_cell = ws.cell(row=row, column=col_offset, value=title)
            title_cell.font = Font(bold=True)
            row += 1

            for col_idx, header in enumerate(df_table.columns, start=col_offset):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.border = border

            for _, record in df_table.iterrows():
                row += 1
                for col_idx, value in enumerate(record, start=col_offset):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.border = border

            return row + 2

        # === Report Layout ===
        row = 2
        row = write_table(row, 2, "📊 Average Grade per Course", avg_grades)
        row = write_table(row, 2, "📉 Average Attendance per Course", avg_attendance)
        row = write_table(row, 2, "📈 Grade Distribution by Course", grade_dist)
        row = write_table(row, 2, "📋 Performance Comparison (Assignments vs Exams)", performance_comparison)
        row = write_table(row, 2, "🚨 At-Risk Students", at_risk)
        row = write_table(row, 2, "🌟 Top Performing Students", top)

        
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for cell in ws[get_column_letter(col_idx)]:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

        with BytesIO() as b:
            wb.save(b)
            return b.getvalue()

    # === CSV Export ===
    elif file_type == "csv":
        with BytesIO() as b:
            def write_section(title, df_table):
                b.write((title + "\n").encode())
                df_table.to_csv(b, index=False)
                b.write(b"\n\n")

            write_section("📊 Average Grade per Course", avg_grades)
            write_section("📉 Average Attendance per Course", avg_attendance)
            write_section("📈 Grade Distribution by Course", grade_dist)
            write_section("📋 Performance Comparison", performance_comparison)
            write_section("🚨 At-Risk Students", at_risk)
            write_section("🌟 Top Performing Students", top)

            return b.getvalue()

    return None

@app.route('/download_excel_students_overview_dashboard')
@login_required
def download_excel_students_overview_dashboard():
    excel_data = generate_students_overview_report("xlsx")
    return send_file(
        BytesIO(excel_data),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='Students_Overview_Report.xlsx'
    )

@app.route('/download_csv_students_overview_dashboard')
@login_required
def download_csv_students_overview_dashboard():
    csv_data = generate_students_overview_report("csv")
    return send_file(
        BytesIO(csv_data),
        mimetype='text/csv',
        as_attachment=True,
        download_name='Students_Overview_Report.csv'
    )




def generate_student_attendance_insights_report(file_type="xlsx"):
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    DATA_DIR = os.path.join(BASE_DIR, "data")

    def load_students(file_name):
        try:
            with open(os.path.join(DATA_DIR, file_name), "r", encoding="utf-8") as f:
                data = json.load(f)
            flat = []
            for entry in data:
                if isinstance(entry, list):
                    flat.extend(entry)
                elif isinstance(entry, dict):
                    flat.append(entry)
            return [(s["ID"], s["Student"], s["Attendance"]) for s in flat if "ID" in s and "Student" in s and "Attendance" in s]
        except:
            return []

    se_students = load_students("software_engineering.json")
    ds_students = load_students("data_science.json")

    if not se_students or not ds_students:
        return None

    se_lectures = [
        "Application Software", "Databases", "Operating Systems", "Networking",
        "Cybersecurity", "Cloud Computing", "Software Testing", "Artificial Intelligence",
        "Machine Learning", "Web Development", "Mobile Development", "DevOps"
    ]
    ds_lectures = [
        "Data Wrangling", "Big Data", "Data Visualization", "Statistics",
        "Machine Learning", "Deep Learning", "NLP", "AI Ethics",
        "Reinforcement Learning", "Cloud Computing for AI", "Model Deployment", "Data Science Projects"
    ]

    def simulate_attendance(students, lectures):
        week_labels = [f"Week {i+1} ({topic})" for i, topic in enumerate(lectures)]
        total_weeks = len(week_labels)
        data = []

        for student_id, name, overall_attendance in students:
            attended_count = round((overall_attendance / 100) * total_weeks)
            attended_weeks = sorted(random.sample(range(total_weeks), attended_count))
            weekly = [100 if i in attended_weeks else 0 for i in range(total_weeks)]
            data.append([student_id, name] + weekly)

        columns = ["Student ID", "Student Name"] + week_labels
        return pd.DataFrame(data, columns=columns)

    df_se = simulate_attendance(se_students, se_lectures)
    df_ds = simulate_attendance(ds_students, ds_lectures)

    if file_type == "xlsx":
        wb = openpyxl.Workbook()
        ws_se = wb.active
        ws_se.title = "Software Engineering"
        ws_ds = wb.create_sheet("Data Science")

        def write_to_sheet(ws, df):
            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

            for col_idx, col_name in enumerate(df.columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.border = border

            for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border

                    
                    if isinstance(value, (int, float)):
                        if value == 100:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
                        elif value == 0:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red

            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max_length + 2

        write_to_sheet(ws_se, df_se)
        write_to_sheet(ws_ds, df_ds)

        b = BytesIO()
        wb.save(b)
        b.seek(0)
        return b.getvalue()

    elif file_type == "csv":
        output = StringIO()
        output.write("Software Engineering\n")
        df_se.to_csv(output, index=False)
        output.write("\n\nData Science\n")
        df_ds.to_csv(output, index=False)
        return output.getvalue().encode("utf-8")

    return None


@app.route("/download_excel_student_attendance_insights_renamed")
def download_excel_student_attendance_insights_renamed():
    content = generate_student_attendance_insights_report("xlsx")
    if not content:
        return "❌ Export failed", 500

    return send_file(
        BytesIO(content),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        download_name="student_attendance_insights.xlsx",
        as_attachment=True
    )



@app.route("/download_csv_student_attendance_insights_renamed")
def download_csv_student_attendance_insights_renamed():
    content = generate_student_attendance_insights_report("csv")
    if not content:
        return "❌ CSV export failed", 500

    return send_file(
        BytesIO(content),
        mimetype="text/csv",
        download_name="student_attendance_insights.csv",
        as_attachment=True
    )





# Function to generate Course Registers Export
def generate_course_registers_report(file_type="xlsx"):
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(BASE_DIR, "course_registers.json")  

    try:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        print(f"❌ Could not load course_registers.json: {e}")
        return None

    if not data or not isinstance(data, list):
        print("❌ Invalid or empty JSON content")
        return None

    df = pd.DataFrame(data)

    
    df_se = df[df["UNI ID"].str.startswith("SE")]
    df_ds = df[df["UNI ID"].str.startswith("DS")]

    if file_type == "xlsx":
        wb = openpyxl.Workbook()
        ws_se = wb.active
        ws_se.title = "Software Engineering"
        ws_ds = wb.create_sheet("Data Science")

        def write_sheet(ws, df_course):
            header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            border = Border(left=Side(style="thin"), right=Side(style="thin"),
                            top=Side(style="thin"), bottom=Side(style="thin"))

            columns = ["UNI ID", "NAME", "Role"]
            for col_num, header in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.border = border

            for row_idx, row in enumerate(df_course[columns].itertuples(index=False), start=2):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border

            for col in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max_len + 2

        write_sheet(ws_se, df_se)
        write_sheet(ws_ds, df_ds)

        b = BytesIO()
        wb.save(b)
        b.seek(0)
        return b.getvalue()

    elif file_type == "csv":
        output = StringIO()
        output.write("Software Engineering\n")
        df_se.to_csv(output, index=False)
        output.write("\n\nData Science\n")
        df_ds.to_csv(output, index=False)
        return output.getvalue().encode("utf-8")

    return None

@app.route("/download_excel_course_registers")
@login_required
def download_excel_course_registers():
    content = generate_course_registers_report("xlsx")
    if not content:
        return "❌ Excel export failed", 500
    return send_file(
        BytesIO(content),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        download_name="Course_Registers.xlsx",
        as_attachment=True
    )

@app.route("/download_csv_course_registers")
@login_required
def download_csv_course_registers():
    content = generate_course_registers_report("csv")
    if not content:
        return "❌ CSV export failed", 500
    return send_file(
        BytesIO(content),
        mimetype="text/csv",
        download_name="Course_Registers.csv",
        as_attachment=True
    )
# --- End of Lecturer side logic ---



# --- Admin side logic ---
@app.route('/home_admin')
@login_required
def home_admin():
    return render_template('home_admin.html', user=current_user)

@app.route('/admin/courses/modules')
def admin_courses_modules():
    return render_template('courses_modules_admin.html')

@app.route('/admin/courses/data-science')
def view_data_science_modules():
    return render_template('data_science_modules.html')

@app.route('/admin/courses/software-engineering')
def view_software_eng_modules():
    return render_template('software_eng_modules.html')


@app.route('/admin/courses/shared')
def view_shared_modules():
    return render_template('shared_modules.html')



@app.route("/admin/modules/web-systems")
def web_systems_dashboard_page():
    return render_template("web_systems_dashboard.html")


@app.route('/admin/modules/ml')
def ml_dashboard_wrapper():
    return render_template('ml_dashboard_wrapper.html')

@app.route('/admin/modules/agile')
def agile_dashboard_wrapper():
    return render_template('agile_dashboard_wrapper.html')


@app.route("/admin/modules/big-data")
def big_data_module_dashboard():
    return render_template("big-data.html")


@app.route("/admin/modules/software-testing")
def software_testing_dashboard_wrapper():
    return render_template("software_testing_dashboard_wrapper.html")

@app.route("/admin/modules/cloud-software")
def cloud_engineering_dashboard_wrapper():
    return render_template("cloud_engineering_dashboard_wrapper.html")

@app.route("/admin/modules/mlops")
def mlops_dashboard_wrapper():
    return render_template("mlops_dashboard_wrapper.html")

@app.route("/admin/modules/data-ethics")
def data_ethics_dashboard_wrapper():
    return render_template("data_ethics_dashboard_wrapper.html")

@app.route("/admin/students")
def student_list():
    import sqlite3
    conn = sqlite3.connect("courses.db")
    query = """
    SELECT s.id, s.username, c.name AS course_name, c.code AS course_code
    FROM students s
    JOIN enrollments e ON s.id = e.student_id
    JOIN courses c ON e.course_id = c.id
    WHERE c.id IN (3, 4) AND s.role = 'Student'
    LIMIT 950
"""

    df = pd.read_sql_query(query, conn)
    conn.close()
    return render_template("students.html", students=df.to_dict("records"))


@app.route('/admin/student_profile/<student_id>')
def student_profile_wrapped(student_id):
    return render_template('student_profile_wrapper.html', student_id=student_id)

@app.route('/admin/lecturers')
def admin_lecturers():
    conn = sqlite3.connect('courses.db')
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("""
        SELECT lecturer_id, lecturer_name, course_code, module_code, module_week
        FROM lecturer_assignments
        ORDER BY lecturer_name, module_code, module_week
    """)
    rows = cur.fetchall()
    conn.close()

    # Group assignments by lecturer
    lecturers = {}
    for row in rows:
        lid = row['lecturer_id']
        if lid not in lecturers:
            lecturers[lid] = {
                'id': row['lecturer_id'],
                'name': row['lecturer_name'],
                'course': row['course_code'],
                'assignments': []
            }
        lecturers[lid]['assignments'].append({
            'module': row['module_code'],
            'week': row['module_week']
        })

    return render_template('lecturers.html', lecturers=list(lecturers.values()))


@app.route("/admin/lecturer_dashboard/<lecturer_id>")
def admin_lecturer_dashboard(lecturer_id):
    import sqlite3
    import random

    module_meta = {
        "SE201": {"weeks": ["Sprint Planning", "Daily Standups", "Backlog Grooming", "Scrum Events", "Velocity Tracking", "Agile Metrics", "Burndown Charts", "Product Increments", "Retrospectives", "Agile Estimation", "Kanban vs Scrum", "Agile Wrap-up"]},
        "SE202": {"weeks": ["HTML Basics", "CSS Styling", "Responsive Design", "JavaScript DOM", "Forms and Validation", "Web Hosting", "REST APIs", "Frontend Frameworks", "Authentication", "Web Security", "Debugging Tools", "Deployment"]},
        "SE203": {"weeks": ["Testing Basics", "Unit Tests", "Mocks and Stubs", "Integration Testing", "System Testing", "Acceptance Testing", "Test Automation", "Bug Tracking", "Regression Testing", "Performance Testing", "Security Testing", "Test Reporting"]},
        "SE204": {"weeks": ["Cloud Basics", "IaaS & PaaS", "Deployment Models", "Cloud Storage", "Load Balancing", "Auto-scaling", "Monitoring Tools", "CI/CD Pipelines", "Containers", "Security in Cloud", "Cloud Costing", "Capstone Demo"]},
        "DS101": {"weeks": ["Data Cleaning", "Feature Engineering", "Model Selection", "Supervised Learning", "Unsupervised Learning", "Neural Networks", "Evaluation Metrics", "Model Deployment", "Overfitting & Underfitting", "Hyperparameter Tuning", "Bias-Variance Tradeoff", "Final Review"]},
        "DS102": {"weeks": ["Intro to Big Data", "Hadoop Ecosystem", "Spark Basics", "Data Lakes & Warehouses", "Data Ingestion", "ETL Pipelines", "MapReduce", "Stream Processing", "Data Storage", "Scalability", "Big Data Tools", "Case Study"]},
        "DS203": {"weeks": ["DevOps & MLOps", "Model Deployment", "API Integration", "Continuous Delivery", "Dockerization", "Monitoring Models", "Data Drift", "Model Logging", "Scaling Inference", "Deployment Tools", "Model Governance", "Final Review"]},
        "DS204": {"weeks": ["Data Ethics Intro", "Bias in AI", "Fairness Metrics", "Case Studies", "Consent Mechanisms", "GDPR", "Data Security", "Responsible AI", "Transparency", "Accountability", "Audit Frameworks", "Wrap-Up"]}
    }

    conn = sqlite3.connect("courses.db")
    cursor = conn.cursor()

    cursor.execute("""
        SELECT DISTINCT lecturer_id, lecturer_name, course_code
        FROM lecturer_assignments
        WHERE lecturer_id = ?
    """, (lecturer_id,))
    row = cursor.fetchone()

    if not row:
        conn.close()
        return "Lecturer not found", 404

    lecturer_id, lecturer_name, course_code = row

    cursor.execute("""
        SELECT module_code, module_week
        FROM lecturer_assignments
        WHERE lecturer_id = ?
    """, (lecturer_id,))
    assignments = cursor.fetchall()
    conn.close()

    week_data = []
    for mod, week in assignments:
        weeks_list = module_meta.get(mod, {}).get("weeks", [])
        try:
            week_index = weeks_list.index(week) + 1
        except ValueError:
            week_index = 999  

        total_students = random.randint(180, 260)
        attended = random.randint(140, total_students)

        week_data.append({
            "module": mod,
            "week": f"{week} (Week {week_index})",
            "week_label": f"{week} (Week {week_index})",
            "week_num": week_index,
            "attended": attended,
            "total": total_students,
            "percentage": round((attended / total_students) * 100, 1)
        })

    
    week_data.sort(key=lambda x: x["week_num"])

    return render_template("admin_lecturer_dashboard.html",
                           lecturer_name=lecturer_name,
                           lecturer_id=lecturer_id,
                           course_code=course_code,
                           week_data=week_data)





 
@app.route('/students')
def students():
    return render_template('students.html')  

@app.route('/reports')
def reports():
    return render_template('reports.html')  


# Database Connection
# ====================
def get_db_connection():
    conn = sqlite3.connect('courses.db')
    conn.row_factory = sqlite3.Row
    return conn


# API: Fetch All Courses
# ====================
@app.route('/api/courses', methods=['GET'])
def get_courses():
    courses = fetch_course_data()  # ✅ Use your cleaned final_data
    return jsonify(courses)


def fetch_course_data():
    conn = sqlite3.connect('courses.db')
    conn.row_factory = sqlite3.Row
    courses = conn.execute('SELECT * FROM courses').fetchall()
    final_data = []

    for course in courses:
        course_id = course['id']

       
        student_count = conn.execute('''
            SELECT COUNT(*) FROM students s
            JOIN enrollments e ON s.id = e.student_id
            WHERE e.course_id = ? AND s.role = 'Student' AND s.enrollment_status != 'Removed'
        ''', (course_id,)).fetchone()[0]

        lecturer_count = conn.execute('''
            SELECT COUNT(*) FROM students s
            JOIN enrollments e ON s.id = e.student_id
            WHERE e.course_id = ? AND s.role = 'Lecturer' AND s.enrollment_status != 'Removed'
        ''', (course_id,)).fetchone()[0]

        final_data.append({
            'id': course_id,
            'name': course['name'],
            'code': course['code'],
            'year': course['year'],   
            'status': course['status'],
            'students': student_count,
            'lecturers': lecturer_count
        })

    conn.close()
    return final_data



# Route: Course Management Page
# ====================
@app.route('/course_management')
def course_management():
    selected_year = request.args.get('year', default='2025')  # default to 2025 if no year selected
    conn = get_db_connection()

    courses = conn.execute('SELECT * FROM courses WHERE year = ?', (selected_year,)).fetchall()
    final_data = []

    for course in courses:
        course_id = course['id']

        student_count = conn.execute('''
            SELECT COUNT(*) FROM students s
            JOIN enrollments e ON s.id = e.student_id
            WHERE e.course_id = ? AND s.role = 'Student' AND s.enrollment_status != 'Removed'
        ''', (course_id,)).fetchone()[0]

        lecturer_count = conn.execute('''
            SELECT COUNT(*) FROM students s
            JOIN enrollments e ON s.id = e.student_id
            WHERE e.course_id = ? AND s.role = 'Lecturer' AND s.enrollment_status != 'Removed'
        ''', (course_id,)).fetchone()[0]

        final_data.append({
            'id': course['id'],
            'name': course['name'],
            'code': course['code'],
            'year': course['year'],
            'status': course['status'],
            'students': student_count,
            'lecturers': lecturer_count
        })

    conn.close()
    return render_template('course_management.html', courses=final_data, selected_year=selected_year)





# API: Fetch all students in a course

@app.route('/api/course/<int:course_id>/students', methods=['GET'])
def get_course_students(course_id):
    conn = get_db_connection()
    students = conn.execute('''
        SELECT s.* FROM students s
        JOIN enrollments e ON s.id = e.student_id
        WHERE e.course_id = ?
    ''', (course_id,)).fetchall()
    conn.close()
    return jsonify([dict(student) for student in students])


# Route: View Course Page (students with 'Removed' status hidden)
# ====================
@app.route('/course/<int:course_id>/view')
def view_course(course_id):
    conn = get_db_connection()

    
    course = conn.execute('SELECT * FROM courses WHERE id = ?', (course_id,)).fetchone()
    if not course:
        conn.close()
        return "Course Not Found", 404

   
    users_in_course = conn.execute('''
    SELECT s.username, s.email, s.role, s.join_date, s.enrollment_status, s.id AS uni_id
    FROM students s
    JOIN enrollments e ON s.id = e.student_id
    WHERE e.course_id = ? AND s.enrollment_status != 'Removed'
    ORDER BY 
        CASE s.role
            WHEN 'Student' THEN 1
            WHEN 'Lecturer' THEN 2
            WHEN 'Teacher Assistant' THEN 3
            WHEN 'Admin' THEN 4
            ELSE 5
        END,
        s.join_date DESC
''', (course_id,)).fetchall()


    conn.close()

    course = dict(course)
    users = [dict(user) for user in users_in_course]

   
    if "Software" in course["name"]:
        template = 'view_software.html'
    elif "Data Science" in course["name"]:
        template = 'view_data_science.html'
    else:
        template = 'view_generic.html'

    print(f"✅ View Course '{course['name']}' — {len(users)} users found")
    if users:
        print(f"Sample user: {users[0]}")

    return render_template(template, course=course, users=users)



# Route: Edit Course Page 
# ====================
@app.route('/course/<int:course_id>/edit', methods=['GET', 'POST'])
def edit_course(course_id):
    conn = get_db_connection()

    if request.method == 'POST':
        name = request.form['name']
        code = request.form['code']
        students = request.form['students']
        lecturers = request.form['lecturers']
        status = request.form['status']
        year = request.form['year']   # 🆕 Added this

        conn.execute('''
    UPDATE courses
    SET name = ?, code = ?, students = ?, lecturers = ?, status = ?, year = ?
    WHERE id = ?
''', (name, code, students, lecturers, status, year, course_id))

        conn.commit()
        conn.close()

        return redirect(url_for('edit_course', course_id=course_id, success='Course updated successfully!'))

    
    course = conn.execute('SELECT * FROM courses WHERE id = ?', (course_id,)).fetchone()

    users_in_course = conn.execute('''
    SELECT s.id, s.username, s.email, s.role, s.join_date, s.enrollment_status, s.id AS uni_id
    FROM students s
    JOIN enrollments e ON s.id = e.student_id
    WHERE e.course_id = ?
    ORDER BY 
        CASE s.role
            WHEN 'Student' THEN 1
            WHEN 'Lecturer' THEN 2
            WHEN 'Teacher Assistant' THEN 3
            WHEN 'Admin' THEN 4
            ELSE 5
        END,
        s.join_date DESC
''', (course_id,)).fetchall()


    conn.close()

    if not course:
        return "Course Not Found", 404

    course = dict(course)
    users = [
    dict(user) | {
        'id': dict(user).get('id') or dict(user).get('uni_id') or 'N/A',
        'uni_id': dict(user).get('uni_id') or dict(user).get('id') or 'N/A'
    }
    for user in users_in_course
]






    if "Software" in course["name"]:
        template = 'edit_software.html'
    elif "Data Science" in course["name"]:
        template = 'edit_data_science.html'
    else:
        template = 'edit_generic.html'

    return render_template(template, course=course, users=users, success=request.args.get('success'))




# Route: Add New User to Course
# ====================
@app.route('/course/<int:course_id>/add_user', methods=['POST'])
def add_user(course_id):
    name = request.form['name']
    role = request.form['role']
    enrollment_status = request.form['enrollment_status']
    course_name = request.form['course_name']
    join_date = datetime.now().strftime('%Y-%m-%d')

    # Generate initials for email
    initials = ''.join([part[0] for part in name.split()]).upper()
    random_number = random.randint(100000, 999999)

    # Generate ID and email
    if role == 'Student':
        uni_id = f"SE{random_number}" if "Software" in course_name else f"DS{random_number}"
        email = f"{initials}{uni_id}@uni.com"
    else:
        uni_id = str(random_number)
        email = f"{name.replace(' ', '').lower()}@uni.com"

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute('''
        INSERT INTO students (id, username, email, role, join_date, enrollment_status)
        VALUES (?, ?, ?, ?, ?, ?)
    ''', (uni_id, name, email, role, join_date, enrollment_status))

   
    student_id = uni_id

    cursor.execute('''
        INSERT INTO enrollments (student_id, course_id)
        VALUES (?, ?)
    ''', (student_id, course_id))

    conn.commit()
    conn.close()

    return redirect(url_for('edit_course', course_id=course_id, success='Person added successfully!'))



# Route: Archive 
# ====================
@app.route('/course/<int:course_id>/delete_user/<user_id>', methods=['POST'])
def delete_user(course_id, user_id):
    conn = get_db_connection()
    conn.execute('''
        UPDATE students
        SET enrollment_status = 'Removed'
        WHERE id = ?
    ''', (user_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('edit_course', course_id=course_id, success='User archived successfully!'))



# Route: Restore Archived User
# ====================
@app.route('/course/<int:course_id>/restore_user/<user_id>', methods=['POST'])
def restore_user(course_id, user_id):
    original_status = request.form.get('original_status', 'Active')

    conn = get_db_connection()
    conn.execute('''
        UPDATE students
        SET enrollment_status = ?
        WHERE id = ?
    ''', (original_status, user_id))
    conn.commit()
    conn.close()

    return redirect(url_for('edit_course', course_id=course_id, success='User restored successfully!'))



# courses/modules help
# ====================
@app.route('/seed_dashboard_data')
def seed_dashboard_data():
    import sqlite3
    import random

    conn = sqlite3.connect("courses.db")
    cursor = conn.cursor()

    
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS dashboard_big_data (
        id TEXT PRIMARY KEY,
        name TEXT,
        assignment_score INTEGER,
        attendance INTEGER
    )
    ''')
    
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS dashboard_web_systems (
        id TEXT PRIMARY KEY,
        name TEXT,
        assignment_score INTEGER,
        attendance INTEGER
    )
    ''')

    
    ds_students = cursor.execute('''
    SELECT s.id, s.username
    FROM students s
    JOIN enrollments e ON s.id = e.student_id
    JOIN courses c ON e.course_id = c.id
    WHERE c.name LIKE '%Data Science%' AND c.year = 2025
    ''').fetchall()

    
    se_students = cursor.execute('''
    SELECT s.id, s.username
    FROM students s
    JOIN enrollments e ON s.id = e.student_id
    JOIN courses c ON e.course_id = c.id
    WHERE c.name LIKE '%Software Engineering%' AND c.year = 2025
    ''').fetchall()

    
    for student_id, name in ds_students:
        score = random.randint(55, 95)
        attendance = random.randint(65, 100)
        cursor.execute('''
        INSERT OR REPLACE INTO dashboard_big_data (id, name, assignment_score, attendance)
        VALUES (?, ?, ?, ?)
        ''', (student_id, name, score, attendance))

    
    for student_id, name in se_students:
        score = random.randint(55, 95)
        attendance = random.randint(65, 100)
        cursor.execute('''
        INSERT OR REPLACE INTO dashboard_web_systems (id, name, assignment_score, attendance)
        VALUES (?, ?, ?, ?)
        ''', (student_id, name, score, attendance))

    conn.commit()
    conn.close()

    return "✅ Dashboard data seeded successfully!"



@app.route('/admin/export_student/<student_id>')
def export_student_data(student_id):
    import pandas as pd
    from io import BytesIO
    import sqlite3
    from student_profile_dashboard import module_meta, course_modules

    conn = sqlite3.connect('courses.db')
    student_query = """
    SELECT s.*, c.name AS course_name, c.code AS course_code
    FROM students s
    JOIN enrollments e ON s.id = e.student_id
    JOIN courses c ON c.id = e.course_id
    WHERE s.id = ?
    """
    df = pd.read_sql(student_query, conn, params=(student_id,))
    conn.close()

    if df.empty:
        return "Student not found."

    student = df.iloc[0]
    username = student["username"]
    email = student["email"]
    course = student["course_name"]
    attendance = student["attendance"]
    course_code = student["course_code"]

    
    data_rows = []
    modules = course_modules.get(course_code, [])

    for mod_code in modules:
        meta = module_meta.get(mod_code)
        if not meta:
            continue

        module_name = meta["name"]
        assignments = meta["assignments"]
        exam_title = meta["exam"]

        for i, assignment in enumerate(assignments):
            key = f"a{i+1}"
            score = student.get(f"{key}_score")
            penalty = student.get(f"{key}_penalty") or "None"
            status = student.get(f"{key}_status") or "Not Completed"

            data_rows.append({
                "Module": f"{module_name} ({mod_code})",
                "Type": f"Assignment {i+1} – {assignment}",
                "Score": score if score is not None else "",
                "Penalty": penalty,
                "Status": status
            })

       
        if exam_title:
            exam_score = student.get("exam_score")
            exam_status = student.get("exam_status") or "Not Completed"
            score = exam_score if exam_status == "Fit to Sit" else ""
            data_rows.append({
                "Module": f"{module_name} ({mod_code})",
                "Type": f"Exam – {exam_title}",
                "Score": score,
                "Penalty": "",
                "Status": exam_status
            })

        # Final Grade
        a1 = float(student.get("a1_score") or 0)
        a2 = float(student.get("a2_score") or 0)
        exam = float(student.get("exam_score") or 0)
        final_grade = round((a1 * 0.25 + a2 * 0.25 + exam * 0.5), 2) if exam_title else round((a1 * 0.5 + a2 * 0.5), 2)

        data_rows.append({
            "Module": f"{module_name} ({mod_code})",
            "Type": "Final Grade",
            "Score": final_grade,
            "Penalty": "",
            "Status": ""
        })

    df_export = pd.DataFrame(data_rows)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Summary sheet
        summary_df = pd.DataFrame([
            ["Name", username],
            ["Email", email],
            ["Course", course],
            ["Attendance (%)", attendance]
        ], columns=["Field", "Value"])
        summary_df.to_excel(writer, index=False, sheet_name="Summary")

        # Performance sheet
        df_export.to_excel(writer, index=False, sheet_name="Performance")

    output.seek(0)
    filename = f"{username}_performance.xlsx"
    return send_file(output, download_name=filename, as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/admin/export_student/<student_id>/pdf')
def export_student_pdf(student_id):
    conn = sqlite3.connect("courses.db")
    cursor = conn.cursor()

    cursor.execute("""
        SELECT s.*, c.name AS course_name, c.code AS course_code
        FROM students s
        JOIN enrollments e ON s.id = e.student_id
        JOIN courses c ON c.id = e.course_id
        WHERE s.id = ?
    """, (student_id,))
    student = cursor.fetchone()
    if not student:
        return "Student not found", 404

    student = dict(zip([column[0] for column in cursor.description], student))
    student_name = student["username"]
    course_code = student["course_code"]
    course = student["course_name"]
    attendance = student["attendance"]

    rows = []
    for mod_code in course_modules.get(course_code, []):
        meta = module_meta.get(mod_code)
        if not meta:
            continue

        module_name = meta["name"]
        assignments = meta["assignments"]
        exam_title = meta["exam"]

        for i, title in enumerate(assignments):
            key = f"a{i+1}"
            score = student.get(f"{key}_score")
            penalty = student.get(f"{key}_penalty") or "None"
            status = student.get(f"{key}_status") or "Not Completed"
            rows.append([
                f"{module_name} ({mod_code})",
                f"Assignment {i+1} - {title}",
                str(score if score is not None else ""),
                penalty,
                status
            ])

        if exam_title:
            exam_score = student.get("exam_score")
            exam_status = student.get("exam_status") or "Not Completed"
            score = exam_score if exam_status == "Fit to Sit" else ""
            rows.append([
                f"{module_name} ({mod_code})",
                f"Exam - {exam_title}",
                str(score),
                "",
                exam_status
            ])

        a1 = float(student.get("a1_score") or 0)
        a2 = float(student.get("a2_score") or 0)
        exam = float(student.get("exam_score") or 0)
        final_grade = round((a1 * 0.25 + a2 * 0.25 + exam * 0.5), 2) if exam_title else round((a1 * 0.5 + a2 * 0.5), 2)
        rows.append([
            f"{module_name} ({mod_code})",
            "Final Grade",
            str(final_grade),
            "",
            ""
        ])

    class PDF(FPDF):
        def header(self):
            self.set_font("Arial", size=12)
            self.cell(0, 10, "Student Report", ln=True, align='C')
            self.ln(5)
            self.set_font("Arial", size=11)
            self.cell(0, 8, f"Name: {student_name}", ln=True)
            self.cell(0, 8, f"ID: {student_id}", ln=True)
            self.cell(0, 8, f"Course: {course}", ln=True)
            self.cell(0, 8, f"Attendance: {attendance}%", ln=True)
            self.ln(10)

    pdf = PDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    headers = ["Module", "Type", "Score", "Penalty", "Status"]
    col_widths = [55, 60, 20, 30, 30]
    line_height = 8

    pdf.set_font("Arial", 'B', 11)
    for i, header in enumerate(headers):
        pdf.cell(col_widths[i], line_height + 2, header, border=1, align='C')
    pdf.ln()

    pdf.set_font("Arial", size=10)

    def calculate_max_lines(texts, widths):
        max_lines = 1
        for i, text in enumerate(texts):
            encoded = str(text).encode('latin-1', errors='replace').decode('latin-1')
            num_lines = pdf.get_string_width(encoded) / (widths[i] - 2)
            max_lines = max(max_lines, int(num_lines) + 1)
        return max_lines

    for row in rows:
        x_start = pdf.get_x()
        y_start = pdf.get_y()

        max_lines = calculate_max_lines(row, col_widths)
        row_height = line_height * max_lines

        for i, item in enumerate(row):
            x = pdf.get_x()
            y = pdf.get_y()
            w = col_widths[i]
            h = row_height
            text = str(item).encode('latin-1', errors='replace').decode('latin-1')

            pdf.multi_cell(w, line_height, text, border=1, align='C')
            pdf.set_xy(x + w, y_start)

        pdf.set_y(y_start + row_height)

    pdf_bytes = pdf.output(dest='S').encode('latin-1', errors='replace')
    buffer = BytesIO(pdf_bytes)
    filename = f"{student_name.replace(' ', '_')}_{student_id}_report.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")



























################# admin lectuer export
@app.route('/admin/export_lecturer/<lecturer_id>')
def export_lecturer_excel(lecturer_id):
    import sqlite3
    import pandas as pd
    import io
    from flask import send_file

    module_meta = {
        "SE201": ["Sprint Planning", "Daily Standups", "Backlog Grooming", "Scrum Events", "Velocity Tracking", "Agile Metrics", "Burndown Charts", "Product Increments", "Retrospectives", "Agile Estimation", "Kanban vs Scrum", "Agile Wrap-up"],
        "SE202": ["HTML Basics", "CSS Styling", "Responsive Design", "JavaScript DOM", "Forms and Validation", "Web Hosting", "REST APIs", "Frontend Frameworks", "Authentication", "Web Security", "Debugging Tools", "Deployment"],
        "SE203": ["Testing Basics", "Unit Tests", "Mocks and Stubs", "Integration Testing", "System Testing", "Acceptance Testing", "Test Automation", "Bug Tracking", "Regression Testing", "Performance Testing", "Security Testing", "Test Reporting"],
        "SE204": ["Cloud Basics", "IaaS & PaaS", "Deployment Models", "Cloud Storage", "Load Balancing", "Auto-scaling", "Monitoring Tools", "CI/CD Pipelines", "Containers", "Security in Cloud", "Cloud Costing", "Capstone Demo"],
        "DS101": ["Data Cleaning", "Feature Engineering", "Model Selection", "Supervised Learning", "Unsupervised Learning", "Neural Networks", "Evaluation Metrics", "Model Deployment", "Overfitting & Underfitting", "Hyperparameter Tuning", "Bias-Variance Tradeoff", "Final Review"],
        "DS102": ["Intro to Big Data", "Hadoop Ecosystem", "Spark Basics", "Data Lakes & Warehouses", "Data Ingestion", "ETL Pipelines", "MapReduce", "Stream Processing", "Data Storage", "Scalability", "Big Data Tools", "Case Study"],
        "DS203": ["DevOps & MLOps", "Model Deployment", "API Integration", "Continuous Delivery", "Dockerization", "Monitoring Models", "Data Drift", "Model Logging", "Scaling Inference", "Deployment Tools", "Model Governance", "Final Review"],
        "DS204": ["Data Ethics Intro", "Bias in AI", "Fairness Metrics", "Case Studies", "Consent Mechanisms", "GDPR", "Data Security", "Responsible AI", "Transparency", "Accountability", "Audit Frameworks", "Wrap-Up"]
    }

    conn = sqlite3.connect("courses.db")
    cursor = conn.cursor()

    cursor.execute("""
        SELECT DISTINCT lecturer_id, lecturer_name, course_code
        FROM lecturer_assignments
        WHERE lecturer_id = ?
    """, (lecturer_id,))
    lecturer = cursor.fetchone()

    if not lecturer:
        return "Lecturer not found", 404

    lecturer_id, lecturer_name, course_code = lecturer

    cursor.execute("""
        SELECT module_code, module_week
        FROM lecturer_assignments
        WHERE lecturer_id = ?
    """, (lecturer_id,))
    assignments = cursor.fetchall()
    conn.close()

    # Prepare attendance data
    rows = []
    for module, week in assignments:
        week_index = None
        if module in module_meta and week in module_meta[module]:
            week_index = module_meta[module].index(week) + 1
        week_full = f"{week} (Week {week_index})" if week_index else week

        total = 250
        attended = 160 + (hash(module + week) % 90)
        percent = round(attended / total * 100, 1)

        rows.append({
            "Module": module,
            "Week": week_full,
            "Students Attended": attended,
            "Total Students": total,
            "Attendance %": percent
        })

    df = pd.DataFrame(rows)

    # Excel export
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Sheet 1: Details
        pd.DataFrame([{
            "Lecturer Name": lecturer_name,
            "Lecturer ID": lecturer_id,
            "Course Code": course_code
        }]).to_excel(writer, sheet_name="Lecturer Info", index=False)

        # Sheet 2: Attendance
        df.to_excel(writer, sheet_name="Attendance", index=False)

    output.seek(0)
    filename = f"{lecturer_name.replace(' ', '_')}_{lecturer_id}.xlsx"
    return send_file(output, as_attachment=True,
                     download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")




@app.route('/admin/export_lecturer/<lecturer_id>/pdf')
def export_lecturer_pdf(lecturer_id):
    import sqlite3
    from fpdf import FPDF
    import io
    from flask import send_file

    module_meta = {
        "SE201": ["Sprint Planning", "Daily Standups", "Backlog Grooming", "Scrum Events", "Velocity Tracking", "Agile Metrics", "Burndown Charts", "Product Increments", "Retrospectives", "Agile Estimation", "Kanban vs Scrum", "Agile Wrap-up"],
        "SE202": ["HTML Basics", "CSS Styling", "Responsive Design", "JavaScript DOM", "Forms and Validation", "Web Hosting", "REST APIs", "Frontend Frameworks", "Authentication", "Web Security", "Debugging Tools", "Deployment"],
        "SE203": ["Testing Basics", "Unit Tests", "Mocks and Stubs", "Integration Testing", "System Testing", "Acceptance Testing", "Test Automation", "Bug Tracking", "Regression Testing", "Performance Testing", "Security Testing", "Test Reporting"],
        "SE204": ["Cloud Basics", "IaaS & PaaS", "Deployment Models", "Cloud Storage", "Load Balancing", "Auto-scaling", "Monitoring Tools", "CI/CD Pipelines", "Containers", "Security in Cloud", "Cloud Costing", "Capstone Demo"],
        "DS101": ["Data Cleaning", "Feature Engineering", "Model Selection", "Supervised Learning", "Unsupervised Learning", "Neural Networks", "Evaluation Metrics", "Model Deployment", "Overfitting & Underfitting", "Hyperparameter Tuning", "Bias-Variance Tradeoff", "Final Review"],
        "DS102": ["Intro to Big Data", "Hadoop Ecosystem", "Spark Basics", "Data Lakes & Warehouses", "Data Ingestion", "ETL Pipelines", "MapReduce", "Stream Processing", "Data Storage", "Scalability", "Big Data Tools", "Case Study"],
        "DS203": ["DevOps & MLOps", "Model Deployment", "API Integration", "Continuous Delivery", "Dockerization", "Monitoring Models", "Data Drift", "Model Logging", "Scaling Inference", "Deployment Tools", "Model Governance", "Final Review"],
        "DS204": ["Data Ethics Intro", "Bias in AI", "Fairness Metrics", "Case Studies", "Consent Mechanisms", "GDPR", "Data Security", "Responsible AI", "Transparency", "Accountability", "Audit Frameworks", "Wrap-Up"]
    }

    conn = sqlite3.connect("courses.db")
    cursor = conn.cursor()

    # Get lecturer info
    cursor.execute("""
        SELECT DISTINCT lecturer_id, lecturer_name, course_code
        FROM lecturer_assignments
        WHERE lecturer_id = ?
    """, (lecturer_id,))
    lecturer = cursor.fetchone()

    if not lecturer:
        return "Lecturer not found", 404

    lecturer_id, lecturer_name, course_code = lecturer

    # Get assignments
    cursor.execute("""
        SELECT module_code, module_week
        FROM lecturer_assignments
        WHERE lecturer_id = ?
    """, (lecturer_id,))
    assignments = cursor.fetchall()
    conn.close()

    # Simulate attendance data
    rows = []
    for module, week in assignments:
        week_index = module_meta.get(module, []).index(week) + 1 if week in module_meta.get(module, []) else '?'
        week_label = f"{week} (Week {week_index})"
        total = 250
        attended = 160 + (hash(module + week) % 90)
        percent = round(attended / total * 100, 1)
        rows.append((module, week_label, attended, total, f"{percent}%"))

    # Generate PDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    pdf.set_title(f"{lecturer_name} ({lecturer_id}) Attendance Report")
    pdf.cell(200, 10, f"Lecturer Report", ln=True, align='C')
    pdf.ln(5)
    pdf.cell(200, 10, f"Name: {lecturer_name}", ln=True)
    pdf.cell(200, 10, f"ID: {lecturer_id}", ln=True)
    pdf.cell(200, 10, f"Course: {course_code}", ln=True)
    pdf.ln(10)

    # Table Header
    pdf.set_font("Arial", 'B', size=11)
    headers = ["Module", "Week", "Attended", "Total", "Attendance %"]
    col_widths = [35, 70, 25, 25, 35]
    for i, header in enumerate(headers):
        pdf.cell(col_widths[i], 10, header, border=1, align='C')
    pdf.ln()

    # Table Body
    pdf.set_font("Arial", size=10)
    for row in rows:
        for i, item in enumerate(row):
            pdf.cell(col_widths[i], 10, str(item), border=1, align='C')
        pdf.ln()

    # Write to actual file-like object
    pdf_bytes = pdf.output(dest='S').encode('latin1')
    buffer = io.BytesIO(pdf_bytes)

    filename = f"{lecturer_name.replace(' ', '_')}_{lecturer_id}_report.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')

#### adding students and lectuers to admin side
@app.route("/api/unassigned_users")
def get_unassigned_users():
    course_id = request.args.get("course_id", type=int)
    role = request.args.get("role", type=str)

    conn = sqlite3.connect("courses.db")
    cur = conn.cursor()

    # Validate course
    cur.execute("SELECT code, name FROM courses WHERE id = ?", (course_id,))
    row = cur.fetchone()
    if not row:
        return jsonify([])

    course_code, course_name = row
    combined = (course_code + course_name).upper()

    prefix = ""
    if "SOFTWARE" in combined or "SE" in combined:
        prefix = "SE"
    elif "DATA" in combined or "DS" in combined:
        prefix = "DS"

    if role.lower() == "student":
        query = """
            SELECT id, username, uni_id FROM students
            WHERE role = 'Student'
              AND enrollment_status != 'Removed'
              AND uni_id IS NOT NULL
              AND id NOT IN (SELECT student_id FROM enrollments)
              AND uni_id LIKE ?
        """
        params = [f"{prefix}%"]

    elif role.lower() == "lecturer":
        query = """
            SELECT id, username, uni_id FROM students
            WHERE role = 'Lecturer'
              AND id NOT IN (SELECT lecturer_id FROM lecturer_assignments)
              AND uni_id IS NOT NULL
        """
        params = []

    else:
        return jsonify([])

    cur.execute(query, params)
    users = [{"id": row[0], "username": row[1], "uni_id": row[2]} for row in cur.fetchall()]
    conn.close()
    return jsonify(users)





@app.route('/admin/add_student_to_course', methods=['POST'])
@login_required
def add_student_to_course():
    data = request.get_json()
    student_id = data.get('student_id')
    course_id = data.get('course_id')

    if not student_id or not course_id:
        return jsonify({'success': False, 'message': 'Missing student_id or course_id'}), 400

    conn = sqlite3.connect('courses.db')
    cursor = conn.cursor()

    
    cursor.execute("SELECT 1 FROM enrollments WHERE student_id = ? AND course_id = ?", (student_id, course_id))
    if cursor.fetchone():
        conn.close()
        return jsonify({'success': False, 'message': 'Student already enrolled'}), 400

   
    cursor.execute("INSERT INTO enrollments (student_id, course_id) VALUES (?, ?)", (student_id, course_id))
    cursor.execute("UPDATE courses SET students = students + 1 WHERE id = ?", (course_id,))
    conn.commit()
    conn.close()

    return jsonify({'success': True, 'message': 'Student added'})


@app.route('/admin/add_lecturer_to_course', methods=['POST'])
@login_required
def add_lecturer_to_course():
    data = request.get_json()
    lecturer_id = data.get('lecturer_id')
    course_id = data.get('course_id')

    if not lecturer_id or not course_id:
        return jsonify({'success': False, 'message': 'Missing lecturer_id or course_id'}), 400

    conn = sqlite3.connect('courses.db')
    cursor = conn.cursor()

    cursor.execute("SELECT code FROM courses WHERE id = ?", (course_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'success': False, 'message': 'Invalid course'}), 404

    course_code = row[0]

    cursor.execute("SELECT 1 FROM lecturer_assignments WHERE lecturer_id = ? AND course_code = ?", (lecturer_id, course_code))
    if cursor.fetchone():
        conn.close()
        return jsonify({'success': False, 'message': 'Lecturer already assigned to this course'}), 400

    
    cursor.execute("SELECT username FROM students WHERE id = ?", (lecturer_id,))
    name_row = cursor.fetchone()
    lecturer_name = name_row[0] if name_row else "Unknown"

    
    cursor.execute("""
        INSERT INTO lecturer_assignments (lecturer_id, lecturer_name, course_code, module_code, module_week)
        VALUES (?, ?, ?, '', '')
    """, (lecturer_id, lecturer_name, course_code))

    
    cursor.execute("INSERT INTO enrollments (student_id, course_id) VALUES (?, ?)", (lecturer_id, course_id))

    
    cursor.execute("UPDATE courses SET lecturers = lecturers + 1 WHERE id = ?", (course_id,))

    conn.commit()
    conn.close()

    return jsonify({'success': True, 'message': 'Lecturer added successfully'})


def get_admin_overview_df():
    return pd.DataFrame(fetch_course_data())

def get_data_science_df():
    return pd.DataFrame(ds_raw)

def get_software_engineering_df():
    return pd.DataFrame(se_raw)

def get_big_data_df():
    return pd.DataFrame(big_data_students)

def get_mlops_df():
    return pd.DataFrame(mlops_raw)

def get_data_ethics_df():
    return pd.DataFrame(ethics_raw)

def get_software_testing_df():
    return pd.DataFrame(testing_raw)

def get_cloud_engineering_df():
    return pd.DataFrame(cloud_raw)

def get_web_systems_df():
    return pd.DataFrame(web_systems_students)



def export_admin_dashboard(file_type, dashboard_name):
    df_fetchers = {
        "overview": get_admin_overview_df,
        "machine_learning": get_data_science_df,
        "agile_development": get_software_engineering_df,
        "big_data": get_big_data_df,
        "mlops": get_mlops_df,
        "data_ethics": get_data_ethics_df,
        "software_testing": get_software_testing_df,
        "cloud_computing": get_cloud_engineering_df,
        "web_systems": get_web_systems_df
    }

    if dashboard_name not in df_fetchers:
        return "Dashboard not found", 404

    df = df_fetchers[dashboard_name]()

    if dashboard_name == "overview" and file_type == "xlsx":
        
        df_main = df.copy()

        df_attendance = df_main[["name", "year", "attendance"]].copy()
        df_attendance.rename(columns={
            "name": "Module",
            "attendance": "Attendance (%)"
        }, inplace=True)

        df_grades = df_main[["name", "year", "average_grade"]].copy()
        df_grades.rename(columns={
            "name": "Module",
            "average_grade": "Grade (%)"
        }, inplace=True)

        wb = Workbook()
        ws_main = wb.active
        ws_main.title = "Course Overview"
        for r in dataframe_to_rows(df_main, index=False, header=True):
            ws_main.append(r)

        ws_att = wb.create_sheet("Avg Attendance")
        for r in dataframe_to_rows(df_attendance, index=False, header=True):
            ws_att.append(r)

        ws_grades = wb.create_sheet("Avg Grades")
        for r in dataframe_to_rows(df_grades, index=False, header=True):
            ws_grades.append(r)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True,
                         download_name="admin_overview.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    #  CSV / Excel
    filename = f"admin_{dashboard_name}.{file_type}"
    if file_type == "xlsx":
        df.to_excel(filename, index=False)
    else:
        df.to_csv(filename, index=False)

    return send_file(filename, as_attachment=True)


# Flask Routes for Each
# -------------------------

@app.route('/download_excel_admin_overview')
@login_required
def download_excel_admin_overview():
    return export_admin_dashboard("xlsx", "overview")

@app.route('/download_csv_admin_overview')
@login_required
def download_csv_admin_overview():
    return export_admin_dashboard("csv", "overview")

@app.route('/download_excel_machine_learning')
@login_required
def download_excel_machine_learning():
    return export_admin_dashboard("xlsx", "machine_learning")

@app.route('/download_csv_machine_learning')
@login_required
def download_csv_machine_learning():
    return export_admin_dashboard("csv", "machine_learning")

@app.route('/download_excel_agile_development')
@login_required
def download_excel_agile_development():
    return export_admin_dashboard("xlsx", "agile_development")

@app.route('/download_csv_agile_development')
@login_required
def download_csv_agile_development():
    return export_admin_dashboard("csv", "agile_development")

@app.route('/download_excel_big_data')
@login_required
def download_excel_big_data():
    return export_admin_dashboard("xlsx", "big_data")

@app.route('/download_csv_big_data')
@login_required
def download_csv_big_data():
    return export_admin_dashboard("csv", "big_data")

@app.route('/download_excel_mlops')
@login_required
def download_excel_mlops():
    return export_admin_dashboard("xlsx", "mlops")

@app.route('/download_csv_mlops')
@login_required
def download_csv_mlops():
    return export_admin_dashboard("csv", "mlops")

@app.route('/download_excel_data_ethics')
@login_required
def download_excel_data_ethics():
    return export_admin_dashboard("xlsx", "data_ethics")

@app.route('/download_csv_data_ethics')
@login_required
def download_csv_data_ethics():
    return export_admin_dashboard("csv", "data_ethics")

@app.route('/download_excel_software_testing')
@login_required
def download_excel_software_testing():
    return export_admin_dashboard("xlsx", "software_testing")

@app.route('/download_csv_software_testing')
@login_required
def download_csv_software_testing():
    return export_admin_dashboard("csv", "software_testing")

@app.route('/download_excel_cloud_computing')
@login_required
def download_excel_cloud_computing():
    return export_admin_dashboard("xlsx", "cloud_computing")

@app.route('/download_csv_cloud_computing')
@login_required
def download_csv_cloud_computing():
    return export_admin_dashboard("csv", "cloud_computing")

@app.route('/download_excel_web_systems')
@login_required
def download_excel_web_systems():
    return export_admin_dashboard("xlsx", "web_systems")

@app.route('/download_csv_web_systems')
@login_required
def download_csv_web_systems():
    return export_admin_dashboard("csv", "web_systems")


@app.route('/admin/export')
@login_required
def admin_export():
    return render_template('admin_export.html')

@app.route('/admin/profile')
@login_required
def admin_profile():
    return render_template('admin_profile.html', user=current_user)


# --- End of Admin logic ---


#  Import Dashboards 
from dashboard_student import init_dashboard as init_student_dashboard
from data_structures_dashboard import init_dashboard as init_dsa_dashboard
from dashboard_web_dev import init_dashboard as init_webdev_dashboard
from dashboard_ai import init_dashboard as init_ai_dashboard
from dashboard_db import init_dashboard as init_db_dashboard
from dashboard_cybersecurity import init_dashboard as init_cybersecurity_dashboard
from dashboard_attendance import init_dashboard as init_attendance_dashboard
#  Import Lecturer Dashboard
from dashboard_lecturer import init_lecturer_dashboard
from software_engineering_dashboard import init_software_engineering_dashboard
from data_science_dashboard import init_data_science_dashboard
from students_overview import init_students_overview
from student_attendance_insights import student_attendance_insights, init_student_attendance_insights
from course_registers import init_course_registers
from dashboard_admin import init_admin_dashboard
from dash_course_dashboard import init_course_dashboard
from dashboard_big_data import init_dashboard as init_big_data_dashboard
from dashboard_web_systems import init_dashboard as init_web_systems_dashboard
from dashboard_professional_practice import init_shared_module_dashboard
from software_testing_dashboard import init_software_testing_dashboard
from cloud_engineering_dashboard import init_cloud_engineering_dashboard
from mlops_dashboard import init_mlops_dashboard
from data_ethics_dashboard import init_data_ethics_dashboard
from student_profile_dashboard import init_student_profile_dashboard



#  Initialize Dashboards
init_student_dashboard(app)
init_dsa_dashboard(app)
init_webdev_dashboard(app)
init_ai_dashboard(app)
init_db_dashboard(app)
init_cybersecurity_dashboard(app)
init_attendance_dashboard(app)
init_lecturer_dashboard(app)  
init_software_engineering_dashboard(app)
init_data_science_dashboard(app)
init_students_overview(app)
app.register_blueprint(student_attendance_insights)
init_student_attendance_insights(app)
init_course_registers(app)
init_admin_dashboard(app)
init_course_dashboard(app)
init_big_data_dashboard(app)
init_web_systems_dashboard(app)
init_shared_module_dashboard(app)
init_software_testing_dashboard(app)
init_cloud_engineering_dashboard(app)
init_mlops_dashboard(app)
init_data_ethics_dashboard(app)
init_student_profile_dashboard(app)

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)
